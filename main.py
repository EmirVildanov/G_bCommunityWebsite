import os
import datetime
import time

import vk
from openpyxl import load_workbook

filename = "LoveTrack.xlsx"
filepath = os.path.join(os.environ["HOMEPATH"], "Desktop", filename)

workbook = load_workbook(filename=filepath)
sheet = workbook.active

minutes_interval_number = 5

# TODO: write check for user online between time intervals
# TODO: transpose table

def find_current_day_row_index():
    current_date_cell_row_index = 2
    current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)
    while current_date_cell.value is not None:
        current_date_cell_row_index += 1
        current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)
    current_date_cell_row_index -= 1
    current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)

    now_date = datetime.datetime.now()
    now_date_truncated = datetime.datetime(now_date.year, now_date.month, now_date.day)

    current_date_cell_value = current_date_cell.value
    if type(current_date_cell_value) != datetime:
        current_date_cell_value = datetime.datetime.strptime(current_date_cell_value, '%d.%m.%Y')

    current_date_truncated = datetime.datetime(current_date_cell_value.year, current_date_cell_value.month,
                                                    current_date_cell_value.day)
    if current_date_truncated == now_date_truncated:
        return current_date_cell_row_index
    else:
        now_date_cell = sheet.cell(row=(current_date_cell_row_index + 1), column=1)
        now_date_cell.value = now_date_truncated.strftime('%d.%m.%Y')
        return current_date_cell_row_index + 1


def write_column_headers():
    current_date_truncated = datetime.datetime(current_date.year, current_date.month, current_date.day)
    minutes_added = datetime.timedelta(minutes=minutes_interval_number)
    for i in range(0, int(24 * 60 / minutes_interval_number)):
        current_cell = sheet.cell(row=1, column=(2 + i))
        current_cell.value = current_date_truncated.strftime('%H:%M')
        current_date_truncated += minutes_added


def check_love_info():
    current_day_row_index = find_current_day_row_index()
    current_minutes_interval = (current_date.time().hour * 60 + current_date.time().minute) // minutes_interval_number

    Milana_id = 96222243
    Egor_id = 82555218
    ids = {"Milana": Milana_id, "Egor": Egor_id}
    cell_string = ""
    for user_name, user_id in ids.items():
        user_info = vk_api.users.get(user_id=user_id, fields="online, last_seen", v="5.131")[0]
        last_seen_info = user_info['last_seen']
        online, last_seen_time, device = user_info['online'], int(last_seen_info['time']), int(last_seen_info['platform'])
        last_seen_time_datetime = (datetime.datetime.utcfromtimestamp(last_seen_time) + datetime.timedelta(hours=5)).strftime('%H:%M')
        print(f"{user_name} online {online}. \t {user_name} last seen {last_seen_time_datetime}")

        # if last item in dictionary
        if user_name == "Egor":
            cell_string += f"{online} - {device}"
        else:
            cell_string += f"{online} - {device} ; "
    sheet.cell(row=current_day_row_index, column=(2 + current_minutes_interval)).value = cell_string
    workbook.save(filename=filepath)
    time.sleep(minutes_interval_number * 60)


if __name__ == "__main__":
    session = vk.Session(access_token='1dd98bfc1dd98bfc1dd98bfc3d1da13be011dd91dd98bfc7d203c7d6d8b1a9c4473cc75')
    vk_api = vk.API(session)
    current_date = datetime.datetime.now()

    while True:
        check_love_info()

# Personal Diagrams view:
# Horizon - time
# Vertical - isOnline. For day -> 0/1. For period -> How often at this time is online

