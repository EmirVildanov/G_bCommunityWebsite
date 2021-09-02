import os
import datetime
import time
from pathlib import Path

import vk
import openpyxl
from openpyxl import load_workbook

from configuration import MINUTES_INTERVAL_NUMBER
from src.constants import ACCESS_TOKEN
from src.online_statusplatform import OnlineStatusPlatform

filename = "LoveTrack.xlsx"
filepath = os.path.join(os.getcwd(), "resources", filename)


# TODO: write check for user online between time intervals
# TODO: transpose table

def find_current_day_row_index(workbook, sheet):
    start_row_index = 2
    current_date_cell_row_index = start_row_index
    current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)

    now_date = datetime.datetime.now()
    now_date_truncated = datetime.datetime(now_date.year, now_date.month, now_date.day)

    if current_date_cell.value is None:
        now_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)
        now_date_cell.value = now_date_truncated.strftime('%d.%m.%Y')
        return start_row_index

    while current_date_cell.value is not None:
        current_date_cell_row_index += 1
        current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)
    current_date_cell_row_index -= 1
    current_date_cell = sheet.cell(row=current_date_cell_row_index, column=1)

    current_date_cell_value = current_date_cell.value
    if type(current_date_cell_value) != datetime:
        current_date_cell_value = datetime.datetime.strptime(current_date_cell_value, '%d.%m.%Y')

    current_date_truncated = datetime.datetime(current_date_cell_value.year, current_date_cell_value.month,
                                               current_date_cell_value.day)
    if current_date_truncated == now_date_truncated:
        return current_date_cell_row_index
    else:
        now_date_cell = sheet.cell(row=(current_date_cell_row_index + 1), column=1)
        now_date_cell.value = now_date_truncated.strftime('%d.%m.%Y')
        return current_date_cell_row_index + 1


def write_column_headers(workbook, sheet):
    current_date = datetime.datetime.now()
    current_date_truncated = datetime.datetime(current_date.year, current_date.month, current_date.day)
    minutes_added = datetime.timedelta(minutes=MINUTES_INTERVAL_NUMBER)
    for i in range(0, int(24 * 60 / MINUTES_INTERVAL_NUMBER)):
        current_cell = sheet.cell(row=1, column=(2 + i))
        current_cell.value = current_date_truncated.strftime('%H:%M')
        current_date_truncated += minutes_added
    workbook.save(filename=filepath)


def check_love_info(workbook, sheet):
    current_day_row_index = find_current_day_row_index(workbook, sheet)
    current_minutes_interval = (current_date.time().hour * 60 + current_date.time().minute) // MINUTES_INTERVAL_NUMBER

    Milana_id = 96222243
    Egor_id = 82555218
    ids = {"Milana": Milana_id, "Egor": Egor_id}
    cell_string = ""
    for user_name, user_id in ids.items():
        user_info = vk_api.users.get(user_id=user_id, fields="online, last_seen", v="5.131")[0]
        last_seen_info = user_info['last_seen']
        online, last_seen_time, device = user_info['online'], int(last_seen_info['time']), int(
            last_seen_info['platform'])
        last_seen_time_datetime = (
                datetime.datetime.utcfromtimestamp(last_seen_time) + datetime.timedelta(hours=5)).strftime('%H:%M')
        print(
            f"{user_name} online {online}. \t {user_name} last seen {last_seen_time_datetime} on {OnlineStatusPlatform(device).name} device")

        # if last item in dictionary
        if user_name == "Egor":
            cell_string += f"{online} - {device}"
        else:
            cell_string += f"{online} - {device} ; "
    sheet.cell(row=current_day_row_index, column=(2 + current_minutes_interval)).value = cell_string
    workbook.save(filename=filepath)


if __name__ == "__main__":
    if Path(filepath).is_file():
        session = vk.Session(access_token=ACCESS_TOKEN)
        vk_api = vk.API(session)
        current_date = datetime.datetime.now()
        workbook = load_workbook(filename=filepath)
        sheet = workbook.active

        while True:
            check_love_info(workbook, sheet)
            time.sleep(MINUTES_INTERVAL_NUMBER * 60)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        write_column_headers(workbook, sheet)
        workbook.save(filepath)

# Personal Diagrams view:
# Horizon - time
# Vertical - isOnline. For day -> 0/1. For period -> How often at this time is online
